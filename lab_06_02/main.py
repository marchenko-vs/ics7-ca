import interface
import openpyxl as xls

TABLE = []


def left_formula(y):
    """1-й столбец"""

    global TABLE

    res = ["-"]

    for i in range(1, len(y)):
        r = y[i] - y[i - 1]
        res.append(str(round(r, 3)))

    TABLE.append(res)


def center_formula(y):
    """2-й столбец"""

    global TABLE

    res = ["-"]

    for i in range(1, len(y) - 1):
        r = (y[i + 1] - y[i - 1]) / 2
        res.append(str(round(r, 3)))
    res.append("-")

    TABLE.append(res)


def equal_param_right_formula(y):
    """3-й столбец"""

    global TABLE

    res = []

    for i in range(len(y) - 2):
        s1 = y[i + 1] - y[i]
        s2 = (y[i + 2] - y[i]) / 2
        r = s1 * 2 - s2
        res.append(str(round(r, 3)))

    res.append("-")
    res.append("-")

    TABLE.append(res)


def var_der_formula(x, y):
    """4-й столбец"""

    global TABLE

    res = []

    for i in range(len(y) - 1):
        s1 = (1 / y[i] - 1 / y[i + 1]) / (1 / x[i] - 1 / x[i + 1])
        r = y[i] * y[i] * s1 / (x[i] * x[i])
        res.append(str(round(r, 3)))

    res.append("-")
    res.append("-")

    TABLE.append(res)


def second_dif_formula(y):
    """5-й столбец"""

    global TABLE

    res = ["-"]
    for i in range(1, len(y) - 1):
        r = y[i - 1] - 2 * y[i] + y[i + 1]
        res.append(str(round(r, 3)))
    res.append("-")

    TABLE.append(res)


def parse_table(name):
    """Загрузка таблицы в программу"""
    try:
        pos = 2
        points = xls.load_workbook(name).active
        table = []

        while points.cell(row=pos, column=1).value is not None:
            table.append([float(points.cell(row=pos, column=1).value),
                          float(points.cell(row=pos, column=2).value)])
            pos += 1

        res = [[], []]

        for i in range(0, len(table)):
            res[0].append(table[i][0])
            res[1].append(table[i][1])

        return res

    except TypeError:
        print("Проверьте данные на входе!")
        return None, None, None

    except ValueError:
        print("Проверьте данные на входе!!!")
        return None, None, None


def main():
    global TABLE

    TABLE = [[1, 2, 3, 4, 5, 6], [0.571, 0.889, 1.091, 1.231, 1.333, 1.412]]

    # 1 столбец -  левосторонняя формула
    left_formula(TABLE[1])

    # 2 столбец - центральная формула
    center_formula(TABLE[1])

    # 3 столбец - вторая формула Рунге с правосторонней формулы
    equal_param_right_formula(TABLE[1])

    # 4 столбец - метод выравнивающих переменных
    var_der_formula(TABLE[0], TABLE[1])

    # 5 столбец - вторая разностная производная
    second_dif_formula(TABLE[1])

    for i in range(6):
        TABLE[0][i] = str(TABLE[0][i])
        TABLE[1][i] = str(TABLE[1][i])

    interface.print_table(TABLE)


if __name__ == "__main__":
    main()
