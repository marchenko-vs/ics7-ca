import openpyxl as xls


def process_table(name):
    try:
        pos = 2
        points = xls.load_workbook(name).active
        table = []

        while points.cell(row=pos, column=1).value is not None:
            table.append([float(points.cell(row=pos, column=1).value),
                          float(points.cell(row=pos, column=2).value)])
            pos += 1

        res = [[], []]

        for i in range(0, len(table)):
            res[0].append(table[i][0])
            res[1].append(table[i][1])

        return res

    except TypeError:
        print("Проверьте данные на входе!")
        return None, None, None

    except ValueError:
        print("Проверьте данные на входе!!!")
        return None, None, None
